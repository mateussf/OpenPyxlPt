from copyreg import constructor
from dataclasses import dataclass
from sqlite3 import Row
import this
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
#from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.borders import Border, Side


import os
import subprocess
import sys


class Planila():
    Arquivo = Workbook()
    ws = ""
    BordaPadrao = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


    def __init__(self, NomeArquivo = ""):
        if (os.path.isfile(NomeArquivo + '.xlsx')):
            this.Arquivo = load_workbook(NomeArquivo + '.xlsx')
        else:
            Arquivo = Workbook()
            Arquivoxlsx = NomeArquivo + '.xlsx'
            Arquivo.save(Arquivoxlsx)
            this.Arquivo = Arquivo

    def CriaAba(self, NomeAba = "", Sobrescrever = True):
        if NomeAba in self.Arquivo.sheetnames:
            ws = self.Arquivo[NomeAba]
            if(Sobrescrever):
                self.LimpaAba(ws)
        else:
            ws = self.Arquivo.create_sheet(NomeAba)

        self.ws = ws


    def DevolveNomeCelula(self, L, I):
        return self.ws.cell(row=L, column=I).coordinate


    def Salvar(self, NomeArquivo, ConverteParaOds = False):
        Arquivoxlsx = NomeArquivo + '.xlsx'
        #Arquivo.save(Arquivoxlsx)

        if "Sheet" in self.Arquivo.sheetnames: #caso exista uma aba sem nome(inicio do arquiv) deleta
            self.Arquivo.remove(self.Arquivo['Sheet'])

        self.Arquivo.save(Arquivoxlsx)
        if (ConverteParaOds):
            self.ConverteXlxsParaOds(Arquivoxlsx)

    def ConverteXlxsParaOds(NomeArquivo):
        #$ /usr/bin/libreoffice --headless --invisible -convert-to ods /home/cwgem/Downloads/QTL_Sample_data.xls
        #convert /home/cwgem/Downloads/QTL_Sample_data.xls -> /home/cwgem/QTL_Sample_data.ods using OpenDocument Spreadsheet Flat XML
        #$ /usr/bin/libreoffice --headless --invisible -convert-to xls /home/cwgem/QTL_Sample_data.ods
        #convert /home/cwgem/QTL_Sample_data.ods -> /home/cwgem/QTL_Sample_data.xls using
        #os.system('soffice --headless --invisible -convert-to ods ' + str(NomeArquivo))

        output = os.popen('soffice --headless --invisible --convert-to ods "' + str(NomeArquivo)).read() +'"'
        #output = os.popen('soffice --convert-to ods ' + str(NomeArquivo) + ' --headless')
        print(output)
        #subprocess.call(["soffice", "--headless", "--invisible", "-convert-to", "ods", NomeArquivo])

    def Fechar(self):
        self.Arquivo.close()

    def DevolveLetraColuna(self, L, I)    :
        return self.ws.cell(row=L, column=I).column_letter

    def AdicionaCabecalho(self, Campos = [], linha = 1, Alinhamento = "", CorFundo = ""):
        Alfabeto = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
        font = Font(name='Arial',
                    size=10,
                    bold=True)

        thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))


        I = 0
        for label in Campos:
            self.ws[Alfabeto[I] + str(linha)] = label
            self.ws[Alfabeto[I] + str(linha)].font = font
            if (Alinhamento != ""):
                self.Alinhar(self.ws, Alfabeto[I] + str(linha), Alinhamento)
            #ws[Alfabeto[I] + '1'].Border = thin_border
            self.ws[Alfabeto[I] + str(linha)].border = thin_border
            if (CorFundo != ""):
                self.ws[Alfabeto[I] + str(linha)].fill = PatternFill("solid", start_color=CorFundo)


            I += 1
        self.ws.append([''])


    def AjustaLarguraColunas(self, PrimeiraLinha = 1):

        TamanhoMinimo = 11
        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                cell.border = self.BordaPadrao
                if cell.row >= PrimeiraLinha:
                    #if column == "G":
                    #    print(len(str(cell.value)))

                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                    adjusted_width = (max_length + 3) * 1.1

                    self.ws.column_dimensions[column].width = adjusted_width

    def AdicionaRegistro(self, Campos = [], Ordenacao = [], DataTypes = [], Borda = False):
        I = 1
        L = self.DevolveUltimaLinha() + 1

        for Conteudo in Campos:
            Celula = self.ws.cell(row=L, column=I)

            if(len(DataTypes) > 0):
                if DataTypes[I-1] == 'numeric':
                    if self.ConverteNumerico(Conteudo):
                        Conteudo = '{0:.2f}'.format(float(Conteudo))

                        Conteudo = float(Conteudo)#str(Conteudo).replace(".", ",")
                    if Conteudo == '0':
                        Conteudo = float('0.00')

                    Celula.number_format = '#,##0.00'
                elif DataTypes[I-1] == "string":
                    Conteudo = str(Conteudo).strip()


            Celula.value = Conteudo

            if(len(Ordenacao) > 0):
                NomeCelula = str(Celula.coordinate)
                self.Alinhar(self.ws,NomeCelula,Ordenacao[I-1], 'center')

            if (Borda):
                Celula.border = self.BordaPadrao

            I += 1


    def LimpaAba(self):
        self.ws.delete_cols(1, 100)
        self.ws.delete_rows(1, 300)

    def FormataCelula(self, cell, formato = ""):
        self.ws[cell].number_format = formato

    def FormataPorcentagem(self, cell):
        self.ws[cell].number_format = '0.00%'

    def MesclarCelulas(self, cells):
        self.ws.merge_cells(cells)

    def Alinhar(self, cell, horizontal='center', vertical='center'):
        currentCell = self.ws[cell]
        currentCell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

    def DevolveUltimaLinha(self, Coluna = ""):
        if Coluna == "":
            return self.ws.max_row
        else:
            return len(self.ws[Coluna])

    def UltimaLinha(self):
        return self.ws.max_row

    def UltimaColuna(self):
        return self.ws.max_column

    def ConverteNumerico(Valor):
        try:
            return float(Valor)
        except:
            return False
            pass

    def AjustaLarguraColuna(self, Coluna, Tamanho = 10):
        self.ws.column_dimensions[Coluna].width = Tamanho

    def AdicionaRegistroUnico(self, Linha = 1, Coluna = 1, Conteudo = '', Negrito = False, SobrePor = True, CorFundo = "", Borda=False):
        '''
            Caso utulizado para formula, deve ser utilizado partindo do principio que está sendo escrito um excel
            Logo, SOMA do BROffice deve ser escrito como SUM que é a formula do excel
        '''
        Celula = self.ws.cell(row=Linha, column=Coluna)
        #print(Celula.value, Conteudo)
        if Celula.value != None:
            if SobrePor:
                Celula.value = Conteudo
        else:
            Celula.value = Conteudo

        if (CorFundo != ""):
            Celula.fill = PatternFill("solid", start_color=CorFundo)

        if (Negrito):
            Celula.font = Font(name='Arial', size=10, bold=True)
        if (Borda):
            Celula.border = self.BordaPadrao

    def AdicionaBorda(self, Linha = 1, Coluna = 1):
        self.ws.cell(row=Linha, column=Coluna).border = self.BordaPadrao